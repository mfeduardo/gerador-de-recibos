from PyQt6 import QtWidgets, uic
from PyQt6.QtWidgets import QMessageBox, QFileDialog
from openpyxl import load_workbook
from copy import copy
import os, time, csv

# váriáveis globais
arquivo = '-'
arquivo_saida = '-'
funcoes_sequecia = []
dados_config = []


def definir_configurcao():
    with open('arquivos/configuracao/config.csv') as configuracao:
        for registro in csv.reader(configuracao):
            global dados_config
            dados_config.append(registro[0])


definir_configurcao()


def alert(msg):
    info = QMessageBox()
    info.setIcon(QMessageBox.Icon.Warning)
    info.setWindowTitle('Atenção!')
    info.setText(f'{msg}         ')
    info.exec()


def open_file():
    global arquivo, arquivo_saida
    file_filter = 'Excel File (*.xlsx *.xls)'
    response = QFileDialog.getOpenFileName(
        filter=file_filter, caption='Selecione a planilha de origem', directory=arquivo)
    if response[0]:
        arquivo = response[0]
        main.lb_file_name.setText(arquivo)
        arquivo_saida, extensao = os.path.splitext(arquivo)
        arquivo_saida += ' - karina.xlsx'
        main.lb_save_name.setText(arquivo_saida)
        main.bt_edit_plan.setEnabled(True)
    else:
        main.lb_file_name.setText(arquivo)
        main.lb_save_name.setText(arquivo_saida)
        if not arquivo:
            main.bt_edit_plan.setEnabled(False)


def save_file():
    global arquivo_saida
    file_filter = 'Excel File (*.xlsx *.xls)'
    response = QFileDialog.getSaveFileName(
        filter=file_filter, caption='Selecione o arquivo', directory=arquivo_saida)
    if response[0]:
        arquivo_saida = response[0]
        arquivo_saida, extensao = os.path.splitext(arquivo_saida)
        arquivo_saida += '.xlsx'
        main.lb_save_name.setText(arquivo_saida)


def processar():
    if (arquivo != '-' and arquivo_saida and funcoes_sequecia):
        main.bt_open.setText('Aguarde...')
        gerar_recibo(arquivo, arquivo_saida)
    else:
        msg = 'Planilha de entrada ou instrução de configuração não informada!'
        alert(msg)


def reset():
    global funcoes_sequecia, arquivo, arquivo_saida
    arquivo = '-'
    arquivo_saida = '-'
    main.lb_save_name.setText(arquivo_saida)
    main.lb_file_name.setText(arquivo_saida)
    main.listViewFunctions.clear()
    main.listViewSheets.clear()
    funcoes_sequecia = []
    main.bt_edit_plan.setEnabled(False)


def selecionar_funcao():
    linha_inicio = main.row_start_input
    linha_final = main.row_end_input
    funcao = main.comboBox.currentIndex()
    legenda = main.comboBox.currentText()

    if funcao >= 0 and (linha_inicio.text().isdigit() and linha_final.text().isdigit()):

        if (int(linha_inicio.text()) > int(linha_final.text())):
            linha_inicio.setFocus()
            return alert('A linha inicial não pode ser maior que a linha final!')

        funcoes_sequecia.append({'funcao': funcao, 'linha_inicio': int(
            linha_inicio.text()), 'linha_final': int(linha_final.text()), 'legenda': legenda})
        main.listViewFunctions.insertItem(len(funcoes_sequecia) - main.listViewFunctions.currentRow(
        ), f'{funcoes_sequecia[main.listViewFunctions.currentRow()][("legenda")]}: {funcoes_sequecia[main.listViewFunctions.currentRow()][("linha_inicio")]} até {funcoes_sequecia[main.listViewFunctions.currentRow()][("linha_final")]}')
        linha_inicio.setText('')
        linha_final.setText('')
        linha_inicio.setFocus()

    else:
        alert('Selecione uma opção e informe todos os campos!')
        linha_inicio.setFocus()


def excluir_funcao():
    funcao_id = main.listViewFunctions.currentRow()

    if funcao_id >= 0:
        msg = QMessageBox()
        msg.setWindowTitle('Pesadelo Karina | Excluir')
        msg.setText('Tem certeza que deseja excluir essa instrução?')
        msg.setIcon(QMessageBox.Icon.Question)
        msg.setStandardButtons(QMessageBox.StandardButton.Yes |
                               QMessageBox.StandardButton.No)
        msg.button(QMessageBox.StandardButton.Yes).setText('Sim')
        msg.button(QMessageBox.StandardButton.No).setText('Não')
        msg.setInformativeText(
            f'{main.listViewFunctions.item(funcao_id).text()}')

        dialogo = msg.exec()

        if dialogo == QMessageBox.StandardButton.Yes:
            del (funcoes_sequecia[funcao_id])
            main.listViewFunctions.takeItem(funcao_id)
            main.row_start_input.setFocus()


def editar_configuracao():
    edit.le_aba_inicial.setText(dados_config[0])
    edit.le_aba_final.setText(dados_config[1])
    edit.le_col_inicial.setText(dados_config[2])
    edit.le_col_final.setText(dados_config[3])
    edit.le_linha_inicial.setText(dados_config[4])
    edit.le_linha_final.setText(dados_config[5])
    edit.le_linha_del.setText(dados_config[6])
    edit.show()


def salvar_configuracao():
    definicoes = [
        edit.le_aba_inicial.text(),
        edit.le_aba_final.text(),
        edit.le_col_inicial.text(),
        edit.le_col_final.text(),
        edit.le_linha_inicial.text(),
        edit.le_linha_final.text(),
        edit.le_linha_del.text()
    ]

    check = True
    for d in definicoes:
        if not d.isdigit():
            check = False

    if check:

        with open('arquivos/configuracao/config.csv', 'w') as configuracao:
            for registro in definicoes:
                print(f'{registro},', file=configuracao)

        global dados_config
        dados_config = definicoes
        definir_configurcao()

        edit.close()

    else:
        alert('- Preencha todos os campos!\n- Informe apenas números!')


def close():
    main.close()


# Remover linhas indesejadas
def gerar_recibo(entrada, saida):
    arquivo_entrada = entrada
    arquivo_saida = saida

    file_path = arquivo_entrada
    workbook = load_workbook(filename=file_path)

    # Índice da terceira aba (0-indexed)
    start_sheet_index = int(dados_config[0])
    end_sheet_index = int(dados_config[1])
    if end_sheet_index > len(workbook.sheetnames):
        end_sheet_index = len(workbook.sheetnames)

    col_init = int(dados_config[2])
    col_end = int(dados_config[2])

    # Funções
    # exclui intervalo

    def excluir_linhas(inicio, fim):
        for row in sheet.iter_rows(min_row=inicio, max_row=fim, min_col=col_init, max_col=col_end):
            rows_to_delete.append(copy(row))
        return

    # verifica e exclui linhas zeradas ou nulas

    def excluir_linhas_zeradas(inicio, fim):
        for row in sheet.iter_rows(min_row=inicio, max_row=fim, min_col=col_init, max_col=col_end):
            if row[0].value in (None, 0):
                rows_to_delete.append(copy(row))
        return


    # verifica e divide por dois as linhas válidas e exclui linhas zeradas ou nulas
    def dividir_excluir_zeradas(inicio, fim):
        for row in sheet.iter_rows(min_row=inicio, max_row=fim, min_col=col_init, max_col=col_end):
            if row[0].value in (None, 0):
                rows_to_delete.append(copy(row))
            else:
                row[0].value = row[0].value/2
        return


    # verifica e multiplica por dois as linhas válidas e exclui linhas zeradas ou nulas
    def multiplicar_excluir_zeradas(inicio, fim):
        for row in sheet.iter_rows(min_row=inicio, max_row=fim, min_col=col_init, max_col=col_end):
            if row[0].value in (None, 0):
                rows_to_delete.append(copy(row))
            else:
                row[0].value = row[0].value*2
        return

    # remover espaços
    colunas = [1, 2, 3, 4, 5, 6, 7, 8]

    def remover_espacos(inicio, fim):
        for coluna in colunas:
            for row in sheet.iter_rows(min_row=inicio, max_row=fim, min_col=coluna, max_col=coluna):
                if type(row[0].value) == str:
                    row[0].value = ' '.join(row[0].value.split())
        return

    # Opercões
    main.progressBar.show()
    main.progressBar.setMinimum(start_sheet_index)
    main.progressBar.setMaximum(end_sheet_index)

    for sheet_index in range(start_sheet_index, end_sheet_index):
        sheet = workbook.worksheets[sheet_index]

        time.sleep(0.01)
        main.progressBar.setValue(sheet_index+1)
        main.listViewSheets.insertItems(
            sheet_index, [f'{sheet_index-start_sheet_index+1}. {sheet.title}'])

        # Definir a última linha com conteúdo
        last_row_with_content = 194

        # Remover linhas vazias abaixo da última linha com conteúdo
        if last_row_with_content < sheet.max_row:
            sheet.delete_rows(last_row_with_content + 1,
                              sheet.max_row - last_row_with_content)

        # Criar uma cópia das linhas para evitar alterar a estrutura durante a iteração
        rows_to_delete = []

        for i in funcoes_sequecia:
            if i['funcao'] == 0:
                excluir_linhas(i["linha_inicio"], i["linha_final"])
            elif i['funcao'] == 1:
                excluir_linhas_zeradas(i["linha_inicio"], i["linha_final"])
            elif i['funcao'] == 2:
                dividir_excluir_zeradas(i["linha_inicio"], i["linha_final"])
            elif i['funcao'] == 3:
                multiplicar_excluir_zeradas(
                    i["linha_inicio"], i["linha_final"])
            elif i['funcao'] == 4:
                remover_espacos(
                    i["linha_inicio"], i["linha_final"])

        # Remover as linhas indesejadas
        for row in rows_to_delete:
            sheet.delete_rows(row[0].row, 1)

    workbook.save(arquivo_saida)
    workbook.close()

    main.progressBar.hide()

    main.bt_open.setText('GERAR RECIBOS')

    alert('Recibos Gerados com Sucesso!  ')


# interface
app = QtWidgets.QApplication([])
# main
main = uic.loadUi('arquivos/ui/main.ui')
edit = uic.loadUi('arquivos/ui/edit.ui')
main.comboBox.addItems(['Excluir Linhas', 'Excluir Linhas Zeradas',
                       'Dividir Por Dois E Excluir Linhas Zeradas', 'Multiplicar Por Dois E Excluir Linhas Zeradas', 'Remover Espaços Em Branco'])
main.bt_open.clicked.connect(processar)
main.bt_add.clicked.connect(selecionar_funcao)
main.bt_remove.clicked.connect(excluir_funcao)
main.progressBar.hide()
main.actionOpen.triggered.connect(open_file)
main.bt_open_plan.clicked.connect(open_file)
main.bt_edit_plan.clicked.connect(save_file)
main.bt_config.clicked.connect(editar_configuracao)
main.bt_reset.clicked.connect(reset)
main.bt_close.clicked.connect(close)
edit.bt_save_config.clicked.connect(salvar_configuracao)
main.actionSalvarArquivo.triggered.connect(save_file)
# inicializar
main.show()
app.exec()
